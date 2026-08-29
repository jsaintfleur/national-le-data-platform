"""U.S. Census Bureau connectors: geography, population, and urbanicity.

Audit findings encoded here:

* A Census API key is now mandatory. Keyless data requests return HTTP 302 to
  missing_key.html -- a client that follows redirects gets HTML, not an error.
* ``place`` requires ``in=state:``; there is no national ``for=place:*``. A national
  place pull is 51 requests.
* The 2025 Gazetteer changed format: 2023/2024 are tab-delimited, 2025 is PIPE-delimited
  and adds a GEOIDFQ column. Column sets also differ between summary levels in the same
  year. Parse per-file; never share one schema.
* ``codes2020/national_county2020.txt`` is frozen at the 2020 vintage and still lists the
  eight abolished Connecticut counties. Counties come from the Gazetteer instead.
* Place-level PEP is bulk-CSV only. The Census API's PEP datasets stop at county level and
  at vintage 2023.
"""
from __future__ import annotations

import io
import zipfile
from pathlib import Path

from ..config import STATES, settings
from ..util.http import FetchResult, download, get_json

GAZ_YEAR = 2025
PEP_VINTAGE = 2025
ACS_VINTAGE = 2024


# --- Gazetteer ---------------------------------------------------------------------------

GAZ_FILES = {
    "place": f"{GAZ_YEAR}_Gaz_place_national.zip",
    "counties": f"{GAZ_YEAR}_Gaz_counties_national.zip",
    "cousubs": f"{GAZ_YEAR}_Gaz_cousubs_national.zip",
    "ua": f"{GAZ_YEAR}_Gaz_ua_national.zip",
}


def gazetteer_url(layer: str) -> str:
    return (f"{settings.census_files}/geo/docs/maps-data/data/gazetteer/"
            f"{GAZ_YEAR}_Gazetteer/{GAZ_FILES[layer]}")


def fetch_gazetteer(layer: str, dest_dir: Path) -> FetchResult:
    return download(gazetteer_url(layer), dest_dir / GAZ_FILES[layer])


def read_gazetteer(zip_path: Path) -> list[dict]:
    """Read a Gazetteer ZIP. Sniffs the delimiter -- 2025 is pipe, earlier years are tab."""
    import csv as _csv

    with zipfile.ZipFile(zip_path) as zf:
        member = next(n for n in zf.namelist() if n.lower().endswith(".txt"))
        raw = zf.read(member).decode("utf-8-sig", "replace")
    head = raw.split("\n", 1)[0]
    delim = "|" if head.count("|") > head.count("\t") else "\t"
    rows = list(_csv.DictReader(io.StringIO(raw), delimiter=delim))
    # 2024 and earlier right-pad the header with spaces.
    return [{(k or "").strip(): (v.strip() if isinstance(v, str) else v)
             for k, v in r.items()} for r in rows]


# --- Population Estimates Program (bulk CSV; the platform's primary denominator) ---------

PEP_FILES = {
    "place": (f"/programs-surveys/popest/datasets/2020-{PEP_VINTAGE}/cities/totals/"
              f"sub-est{PEP_VINTAGE}.csv"),
    "county": (f"/programs-surveys/popest/datasets/2020-{PEP_VINTAGE}/counties/totals/"
               f"co-est{PEP_VINTAGE}-alldata.csv"),
    "state": (f"/programs-surveys/popest/datasets/2020-{PEP_VINTAGE}/state/totals/"
              f"NST-EST{PEP_VINTAGE}-ALLDATA.csv"),
}


def fetch_pep(level: str, dest_dir: Path) -> FetchResult:
    path = PEP_FILES[level]
    return download(settings.census_files + path, dest_dir / Path(path).name)


def read_pep_csv(path: Path) -> list[dict]:
    import csv as _csv
    text = path.read_text(encoding="latin-1")
    return list(_csv.DictReader(io.StringIO(text)))


# --- ACS (demographic composition; secondary denominator) --------------------------------

def acs5(get_vars: list[str], for_clause: str, in_clause: str | None = None,
         vintage: int = ACS_VINTAGE) -> list[list[str]]:
    if not settings.census_key:
        raise RuntimeError(
            "CENSUS_API_KEY is not set. A key is now mandatory; keyless requests return "
            "HTTP 302 to missing_key.html. Free signup: "
            "https://api.census.gov/data/key_signup.html"
        )
    params = {"get": ",".join(get_vars), "for": for_clause, "key": settings.census_key}
    if in_clause:
        params["in"] = in_clause
    payload = get_json(f"{settings.census_api}/{vintage}/acs/acs5", params)
    return payload  # type: ignore[return-value]


def acs5_places_all_states(get_vars: list[str], vintage: int = ACS_VINTAGE) -> list[dict]:
    """National place pull. 51 requests -- place has no national wildcard."""
    from ..util.fips import STATE_FIPS

    out: list[dict] = []
    for st in STATES:
        fips = STATE_FIPS.get(st)
        if not fips:
            continue
        try:
            rows = acs5(get_vars, "place:*", f"state:{fips}", vintage)
        except Exception:
            continue
        if not rows:
            continue
        header, *body = rows
        out.extend(dict(zip(header, r)) for r in body)
    return out


def acs5_counties_national(get_vars: list[str], vintage: int = ACS_VINTAGE) -> list[dict]:
    rows = acs5(get_vars, "county:*", "state:*", vintage)
    header, *body = rows
    return [dict(zip(header, r)) for r in body]


def acs5_cousubs_state(state_fips: str, get_vars: list[str],
                       vintage: int = ACS_VINTAGE) -> list[dict]:
    rows = acs5(get_vars, "county subdivision:*", f"state:{state_fips}&in=county:*", vintage)
    header, *body = rows
    return [dict(zip(header, r)) for r in body]


# --- Urban areas (urbanicity band) --------------------------------------------------------

UA_LIST_URL = f"{settings.census_files}/geo/docs/reference/ua/2020_Census_ua_list_all.xlsx"
UA_PLACE_XWALK_URL = (f"{settings.census_files}/geo/docs/maps-data/data/rel2020/ua/"
                      "tab20_ua20_place20_natl.txt")
UA_COUNTY_XWALK_URL = (f"{settings.census_files}/geo/docs/maps-data/data/rel2020/ua/"
                       "tab20_ua20_county20_natl.txt")
PLACE_BY_COUNTY_URL = (f"{settings.census_files}/geo/docs/reference/codes2020/"
                       "national_place_by_county2020.txt")
PLACE_CODES_URL = (f"{settings.census_files}/geo/docs/reference/codes2020/"
                   "national_place2020.txt")


def fetch_simple(url: str, dest_dir: Path) -> FetchResult:
    return download(url, dest_dir / Path(url).name)


def read_pipe_table(path: Path) -> list[dict]:
    import csv as _csv
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    return list(_csv.DictReader(io.StringIO(text), delimiter="|"))


def read_ua_list(path: Path) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(it)]
    out = []
    for row in it:
        if row is None or row[0] is None:
            continue
        rec = dict(zip(header, row))
        rec["UACE"] = str(rec.get("UACE", "")).zfill(5)
        out.append(rec)
    wb.close()
    return out
