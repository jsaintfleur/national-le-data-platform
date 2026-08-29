"""Census government-finance connector.

This is the most attribution-hazardous source in the platform, and the connector is written
to make the hazard structural rather than a matter of discipline:

* The unit of observation is a GOVERNMENT UNIT (a city, a county), never a police agency.
  No published crosswalk from a Census government ID to an FBI ORI exists. The platform
  therefore never emits a per-agency spending figure.
* The finance key is 12 characters, whose last 6 are CENSUS_ID_PID6. The 14-character
  CENSUS_ID_GIDID appears in the 2022 Government Units listing and was DROPPED from 2025.
  Never key on GIDID.
* Item code E62 is police-protection current operations. F62 is CONSTRUCTION, not capital
  outlay: G62 (land) and K62 (equipment) have zero records in the public-use file, and K62
  is not collected for most local governments.
* The annual survey is a voluntary stratified SAMPLE. 2024 covers 57.8% of counties, 20.8%
  of municipalities, 4.9% of townships, and 26.5% of E62 values are imputed. Census years
  (ending 2 and 7) are a full universe. A series that crosses that boundary is
  discontinuous by design, not by accident.
* Survey year N spans individual fiscal years ending 1 Jul N-1 through 30 Jun N -- a
  ~24-month window across units. fiscal_year_ending is a first-class column here.
"""
from __future__ import annotations

import io
import zipfile
from pathlib import Path

from ..config import settings
from ..util.http import FetchResult, download

POLICE_ITEM_CODES = {
    "E62": "Police protection - current operations",
    "F62": "Police protection - construction",
    "G62": "Police protection - land and existing structures (absent from public-use file)",
    "K62": "Police protection - equipment (not collected for most local governments)",
    "M62": "Intergovernmental to local governments, police protection",
}
# Stored alongside E62 because a sheriff's E62 excludes the jail and the civil/bailiff
# functions his headcount includes. Without these the number is uninterpretable.
CONTEXT_ITEM_CODES = {
    "E04": "Correctional institutions - current operations",
    "E05": "Other corrections - current operations",
    "E25": "Judicial and legal - current operations",
    "E89": "Other and unallocable - current operations",
}

IMPUTATION_FLAGS = {
    "R": "Reported by the government",
    "I": "Imputed by the Census Bureau",
    "S": "Taken from an alternative source",
    "A": "Analyst correction",
    "M": "Unknown",
    "N": "Not applicable",
}

GOV_TYPE = {
    "0": "State", "1": "County", "2": "City", "3": "Township",
    "4": "Special District", "5": "Independent School District",
}

# Filenames are NOT templatable: 2022 is "_Individual_Unit_File.zip" (singular),
# 2023 and 2024 are "_Individual_Unit_Files.zip" (plural).
IUF_URLS = {
    2024: "/programs-surveys/gov-finances/tables/2024/2024_Individual_Unit_Files.zip",
    2023: "/programs-surveys/gov-finances/tables/2023/2023_Individual_Unit_Files.zip",
    2022: "/programs-surveys/gov-finances/tables/2022/2022_Individual_Unit_File.zip",
}
GUS_URLS = {
    2025: "/programs-surveys/gus/datasets/2025/gov_units_2025.zip",
    2022: "/programs-surveys/gus/datasets/2022/govt_units_2022.ZIP",  # uppercase extension
}


def fetch_iuf(year: int, dest_dir: Path) -> FetchResult:
    return download(settings.census_files + IUF_URLS[year],
                    dest_dir / f"{year}_Individual_Unit_File.zip")


def fetch_gus(year: int, dest_dir: Path) -> FetchResult:
    return download(settings.census_files + GUS_URLS[year], dest_dir / f"gov_units_{year}.zip")


def is_census_of_governments_year(year: int) -> bool:
    """Full-universe canvass years end in 2 or 7."""
    return year % 5 == 2


# --- Fixed-width layouts (validated against downloaded files) ---------------------------

def parse_fin_record(line: str, wanted: set[str] | None = None) -> dict | None:
    """FinEstDAT record: 32 chars. Amounts are in THOUSANDS of dollars.

    Layout verified against the downloaded 2024 file, all 511,362 rows of which are exactly
    32 characters: ID(12) + item code(3) + amount(12) + year(4) + flag(1).
    """
    if len(line) < 32:
        return None
    item = line[12:15].strip()
    if wanted is not None and item not in wanted:
        return None
    amount = line[15:27].strip()
    year = line[27:31].strip()
    gov_id = line[0:12]
    return {
        "census_gov_id_12": gov_id,
        "state_fips": gov_id[0:2],
        "gov_type_code": gov_id[2:3],
        "gov_type": GOV_TYPE.get(gov_id[2:3]),
        "county_fips_within_state": gov_id[3:6],
        "pid6": gov_id[6:12],
        "item_code": item,
        "amount_thousands": int(amount) if amount.lstrip("-").isdigit() else None,
        "survey_year": int(year) if year.isdigit() else None,
        "data_flag": line[31:32].strip() or None,
    }


def parse_pid_record(line: str) -> dict | None:
    """Fin_PID record: 146 chars. Carries the FIPS place crosswalk at 112-116."""
    if len(line) < 146:
        line = line.ljust(146)
    gov_id = line[0:12]
    if not gov_id.strip():
        return None
    pop = line[116:125].strip()
    fye = line[140:144].strip()
    return {
        "census_gov_id_12": gov_id,
        "pid6": gov_id[6:12],
        "unit_name": line[12:76].strip(),
        "county_name": line[76:111].strip(),
        "fips_place": (line[111:116].strip() or None),
        "population": int(pop) if pop.isdigit() else None,
        "population_year": line[125:127].strip() or None,
        "special_district_function": line[136:138].strip() or None,
        "fiscal_year_ending": fye or None,  # MMDD
        "survey_year": line[144:146].strip() or None,
    }


def read_iuf_zip(zip_path: Path, wanted_items: set[str] | None = None
                 ) -> tuple[list[dict], list[dict]]:
    """Return (finance rows, unit-directory rows) from an Individual Unit File ZIP."""
    fin: list[dict] = []
    pid: list[dict] = []
    with zipfile.ZipFile(zip_path) as zf:
        for name in zf.namelist():
            low = name.lower()
            if low.endswith(".pdf"):
                continue
            if "finestdat" in low:
                with zf.open(name) as fh:
                    for raw in io.TextIOWrapper(fh, encoding="latin-1", newline=""):
                        rec = parse_fin_record(raw.rstrip("\r\n"), wanted_items)
                        if rec:
                            fin.append(rec)
            elif low.startswith("fin_pid") or "fin_pid" in low:
                with zf.open(name) as fh:
                    for raw in io.TextIOWrapper(fh, encoding="latin-1", newline=""):
                        rec = parse_pid_record(raw.rstrip("\r\n"))
                        if rec:
                            pid.append(rec)
    return fin, pid


def read_gus_general_purpose(zip_path: Path) -> list[dict]:
    """General-purpose government directory: PID6 plus FIPS_STATE/COUNTY/PLACE."""
    from openpyxl import load_workbook

    with zipfile.ZipFile(zip_path) as zf:
        member = next((n for n in zf.namelist() if n.lower().endswith(".xlsx")), None)
        if member is None:
            return []
        data = zf.read(member)
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = next((s for s in wb.sheetnames if "general" in s.lower()), wb.sheetnames[0])
    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(it)]
    rows: list[dict] = []
    for row in it:
        if row is None or row[0] is None:
            continue
        rec = dict(zip(header, row))
        rec["CENSUS_ID_PID6"] = str(rec.get("CENSUS_ID_PID6", "")).strip().zfill(6)
        for col, width in (("FIPS_STATE", 2), ("FIPS_COUNTY", 3), ("FIPS_PLACE", 5)):
            v = rec.get(col)
            rec[col] = str(v).strip().zfill(width) if v not in (None, "") else None
        rows.append(rec)
    wb.close()
    return rows
